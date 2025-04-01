# 01 => IMPORTACOES DE MODULOS
import sys
import subprocess
import traceback
import shutil
import pandas as pd
import os
import re
import PyQt5.QtWidgets as QtW
import PyQt5.QtCore as QtC
from PyQt5 import QtGui
from openpyxl import load_workbook
from openpyxl.styles import PatternFill
from unidecode import unidecode
from PyQt5 import uic


# 02 => INTERFACE GRAFICA - GUI (Graphical User Interface)
class UI(QtW.QMainWindow):

    def __init__(self):
        super(UI, self).__init__()

        # Carrega o arquivo GUI
        root_folder = os.path.dirname(os.path.abspath(__file__))
        arquivo_gui = os.path.join(root_folder, "GUI_Interface.ui")
        uic.loadUi(arquivo_gui, self)

        # Adiciona um nome a janela
        self.setWindowTitle("PCP Usinagem - Programa Auxiliar")

        # Definicao dos widgets importados no arquivo 'gui'.
        self.tab_Gerador = self.findChild(QtW.QTabWidget, "tab_Gerador")
        self.pushButton_ProcurarPCP = self.findChild(QtW.QPushButton, "pushButton_ProcurarPCP")
        self.pushButton_ProcurarDesenhos = self.findChild(QtW.QPushButton, "pushButton_ProcurarDesenhos")
        self.pushButton_GerarPastas = self.findChild(QtW.QPushButton, "pushButton_GerarPastas")
        self.pushButton_Atualizar = self.findChild(QtW.QPushButton, "pushButton_Atualizar")
        self.pushButton_LimparTudo = self.findChild(QtW.QPushButton, "pushButton_LimparTudo")
        self.pushButton_reserva = self.findChild(QtW.QPushButton, "pushButton_reserva")
        self.radioButton_codigo = self.findChild(QtW.QRadioButton, "radioButton_codigo")
        self.radioButton_referencia = self.findChild(QtW.QRadioButton, "radioButton_referencia")
        self.checkBox_Pintura = self.findChild(QtW.QCheckBox, "checkBox_Pintura")
        self.checkBox_Usinagem = self.findChild(QtW.QCheckBox, "checkBox_Usinagem")
        self.checkBox_Processos = self.findChild(QtW.QCheckBox, "checkBox_Processos")
        self.checkBox_FornUsi = self.findChild(QtW.QCheckBox, "checkBox_FornUsi")
        self.lineEdit_ArquivoPCP = self.findChild(QtW.QLineEdit, "lineEdit_ArquivoPCP")
        self.lineEdit_Desenhos = self.findChild(QtW.QLineEdit, "lineEdit_Desenhos")
        self.tab_ImportarLista = self.findChild(QtW.QTabWidget, "tab_ImportarLista")
        self.pushButton_ArquivoPMS = self.findChild(QtW.QPushButton, "pushButton_ArquivoPMS")
        self.pushButton_ArquivoPCPPadrao = self.findChild(QtW.QPushButton, "pushButton_ArquivoPCPPadrao")
        self.pushButton_GerarPCP = self.findChild(QtW.QPushButton, "pushButton_GerarPCP")
        self.pushButton_VerificarArquivos = self.findChild(QtW.QPushButton, "pushButton_VerificarArquivos")
        self.lineEdit_ArquivoPMS = self.findChild(QtW.QLineEdit, "lineEdit_ArquivoPMS")
        self.lineEdit_ArquivoPCPPadrao = self.findChild(QtW.QLineEdit, "lineEdit_ArquivoPCPPadrao")
        self.lineEdit_NomeProjeto = self.findChild(QtW.QLineEdit, "lineEdit_NomeProjeto")
        self.lineEdit_NomeLista = self.findChild(QtW.QLineEdit, "lineEdit_NomeLista")
        self.StatusPanel = self.findChild(QtW.QPlainTextEdit, "StatusPanel")

        # Acoes dos widgets
        self.pushButton_ProcurarPCP.clicked.connect(self.procurar_arquivo_pcp)
        self.pushButton_ProcurarDesenhos.clicked.connect(self.procurar_pasta_desenhos)
        self.pushButton_GerarPastas.clicked.connect(self.gerar_pastas)
        self.pushButton_Atualizar.clicked.connect(self.atualizar)
        self.pushButton_LimparTudo.clicked.connect(self.start_countdown)
        # self.pushButton_reserva.clicked.connect(self.GERAR_PLANILHAS)
        self.pushButton_ArquivoPMS.clicked.connect(self.procurar_arquivo_csv)
        self.pushButton_ArquivoPCPPadrao.clicked.connect(self.procurar_arquivo_pcp_padrao)
        self.pushButton_GerarPCP.clicked.connect(self.gerar_arquivo_pcp_csv)
        self.pushButton_VerificarArquivos.clicked.connect(self.verificar_arquivos_desenhos_pms)

        self.lineEdit_ArquivoPCP.textChanged.connect(self.liberar_botoes)
        self.lineEdit_Desenhos.textChanged.connect(self.liberar_botoes)
        self.radioButton_codigo.toggled.connect(self.liberar_botoes)
        self.radioButton_referencia.toggled.connect(self.liberar_botoes)
        self.checkBox_Pintura.stateChanged.connect(self.liberar_botoes)
        self.checkBox_Usinagem.stateChanged.connect(self.liberar_botoes)
        self.checkBox_Processos.stateChanged.connect(self.liberar_botoes)
        self.checkBox_FornUsi.stateChanged.connect(self.liberar_botoes)
        self.lineEdit_ArquivoPMS.textChanged.connect(self.liberar_botoes)
        self.lineEdit_ArquivoPCPPadrao.textChanged.connect(self.liberar_botoes)
        self.lineEdit_NomeProjeto.textChanged.connect(self.liberar_botoes)
        self.lineEdit_NomeLista.textChanged.connect(self.liberar_botoes)

        # Redireciona a saída da função Print para o widget StatusPanel
        sys.stdout.write = self.StatusPanel.insertPlainText

        # +++++++++++++++++++++++++++++++++++
        # + DECLARACAO DE VARIAVEIS GLOBAIS +
        # +++++++++++++++++++++++++++++++++++

        self.PCP_Path = self.lineEdit_ArquivoPCP.text()
        self.ListaPCP = None
        self.Desenhos_Path = self.lineEdit_Desenhos.text()
        self.ListaArquivos = []
        self.Arquivo_CSV = None

        # Cria a instância de configurações gerais para uso nas funções
        self.settings = QtC.QSettings()

        # Cria o argumento inicial "Ultima pasta acessada" contendo o caminho do arquivo python
        self.settings.setValue("Ultima Pasta Acessada", os.path.dirname(os.path.abspath(__file__)))

        # Variavel utilizada na função de limpeza
        self.countdown_timer = QtC.QTimer(self)
        self.countdown_timer.timeout.connect(self.update_countdown)
        self.countdown = 5

        # Apresenta a janela do aplicativo
        self.show()

# +++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++

# 04 => FUNCOES

    # Bloqueia/desbloqueia os botoes de acao da tab1
    def liberar_botoes(self):
        l_pcp = self.lineEdit_ArquivoPCP.text()
        l_desenhos = self.lineEdit_Desenhos.text()
        l_pms = self.lineEdit_ArquivoPMS.text()
        l_pcpp = self.lineEdit_ArquivoPCPPadrao.text()
        l_proj = self.lineEdit_NomeProjeto.text()
        l_lista = self.lineEdit_NomeLista.text()
        checkbox_processos = True if self.checkBox_Processos.isChecked() else False
        checkbox_pintura = True if self.checkBox_Pintura.isChecked() else False
        checkbox_usinagem = True if self.checkBox_Usinagem.isChecked() else False
        checkbox_fornusi = True if self.checkBox_FornUsi.isChecked() else False

        if l_pcp:
            self.pushButton_LimparTudo.setEnabled(True)
        else:
            self.pushButton_LimparTudo.setEnabled(False)

        if l_pcp and l_desenhos:
            if checkbox_processos or checkbox_pintura or checkbox_usinagem or checkbox_fornusi:
                self.pushButton_GerarPastas.setEnabled(True)
                self.pushButton_Atualizar.setEnabled(True)
            else:
                self.pushButton_GerarPastas.setEnabled(False)
                self.pushButton_Atualizar.setEnabled(False)

        if l_pms and l_desenhos:
            self.pushButton_VerificarArquivos.setEnabled(True)
        else:
            self.pushButton_VerificarArquivos.setEnabled(False)

        if l_pms and l_pcpp and l_proj and l_lista:
            self.pushButton_GerarPCP.setEnabled(True)
        else:
            self.pushButton_GerarPCP.setEnabled(False)

    # Abre a janela para selecionar o arquivo de PCP
    def procurar_arquivo_pcp(self):
        # Abre a janela de interação com o usario para procurar o arquivo de PCP
        fname1 = QtW.QFileDialog.getOpenFileName(
            parent=self,
            caption="Selecione um arquivo",
            directory=self.settings.value("Ultima Pasta Acessada"),
            filter="Excel Files (*.xls*)"
        )
        # Escreve o caminho do arquivo selecionado na linha correspondente da interface
        if fname1[0]:
            self.lineEdit_ArquivoPCP.setText(str(fname1[0]))
            print("Caminho do arquivo PCP selecionado:", fname1[0])
            ultima_pasta_acessada = os.path.dirname(fname1[0])
            self.settings.setValue("Ultima Pasta Acessada", ultima_pasta_acessada)
        else:
            print("Nenhum arquivo foi selecionado.")
            return None
        return fname1

    # Abre a janela para selecionar a pasta contendo os arquivos das peças
    def procurar_pasta_desenhos(self):
        # Open file dialog
        fname2 = QtW.QFileDialog.getExistingDirectory(
            parent=self,
            caption="Selecione o local dos desenhos",
            directory=self.settings.value("Ultima Pasta Acessada"),
            options=QtW.QFileDialog.ShowDirsOnly
        )
        # Escreve o caminho na linha
        if fname2:
            fname2 = QtC.QDir.cleanPath(fname2)
            self.lineEdit_Desenhos.setText(fname2)
            print("Caminho dos desenhos: ", fname2)
            ultima_pasta_acessada = os.path.dirname(fname2)
            self.settings.setValue("Ultima Pasta Acessada", ultima_pasta_acessada)
        else:
            print("Nenhuma pasta foi selecionada.")
            return None
        return fname2

    # Abre a janela para selecionar o arquivo de CSV com o PMS
    def procurar_arquivo_csv(self):
        # Abre a janela de interação com o usario para procurar o arquivo de PMS da lista
        pmspath = QtW.QFileDialog.getOpenFileName(
            parent=self,
            caption="Abrir Arquivo",
            directory=self.settings.value("Ultima Pasta Acessada"),
            filter="CSV Files (*.csv)"
        )
        # Escreve o caminho do arquivo selecionado na linha correspondente da interface
        if pmspath[0]:
            self.lineEdit_ArquivoPMS.setText(str(pmspath[0]))
            print("Caminho do arquivo PMS(.csv): ", (pmspath[0]))
            ultima_pasta_acessada = os.path.dirname(pmspath[0])
            self.settings.setValue("Ultima Pasta Acessada", ultima_pasta_acessada)
        else:
            print("Nenhum arquivo foi selecionado.")
            return None
        return pmspath

    # Abre a janela para selecionar o arquivo de Excel com o modelo padrão de preenchimento do PCP
    def procurar_arquivo_pcp_padrao(self):
        # Abre janela para selecao do arquivo PCP Padrao vazio
        caminho_pcp_padrao = QtW.QFileDialog.getOpenFileName(
            parent=self,
            caption="Abrir Arquivo",
            directory=self.settings.value("Ultima Pasta Acessada"),
            filter="XLSM Files (*.xlsm)"
        )
        # Excreve o caminho do arquivo selecionado no widget correspondente
        if caminho_pcp_padrao[0]:
            self.lineEdit_ArquivoPCPPadrao.setText(str(caminho_pcp_padrao[0]))
            print("Caminho do arquivo PCP Padrão: ", (caminho_pcp_padrao[0]))
        else:
            print("Nenhum arquivo foi selecionado.")
            return None
        return caminho_pcp_padrao

    # Parte do botao de limpeza
    def start_countdown(self):
        if self.countdown_timer.isActive():
            self.limpeza()  # Chama função que realiza a limpeza
            self.countdown_timer.stop()
            self.pushButton_LimparTudo.setText('Limpar tudo')
            self.countdown = 5
        else:
            self.countdown_timer.start(1000)
            self.pushButton_LimparTudo.setText(f'Confirma? [{self.countdown}]')

    # Parte do botao de limpeza
    def update_countdown(self):
        self.countdown -= 1
        if self.countdown > 0:
            self.pushButton_LimparTudo.setText(f'Confirma? [{self.countdown}]')
        else:
            self.countdown_timer.stop()
            self.pushButton_LimparTudo.setText('Limpar tudo')
            self.countdown = 5

    # Função que cria as variáveis gerais de uso nas demais funções
    def ler_arquivo_pcp(self):
        try:
            # A = CODIGO
            # B = DESCRICAO
            # C = REFERENCIA
            # D = PESO
            # E = QUANTIDADE
            # F = MATERIAL
            # G = PROCESSOS
            # H = DIMENSIONAMENTO MATERIA PRIMA
            # I = TRATAMENTO TERMICO
            # J = ACABAMENTO SUPERFICIAL
            # K = PORTE USINAGEM
            # M = FORN. USINAGEM

            # Carrega o arquivo Excel
            self.PCP_Path = self.lineEdit_ArquivoPCP.text()
            # noinspection PyTypeChecker
            self.ListaPCP = pd.read_excel(
                io=self.PCP_Path,
                sheet_name="Principal",
                usecols="A,B,C,D,E,F,G,H,I,J,K,M",
                skipfooter=1,
                dtype=str
            )

            self.ListaPCP.columns = ['A', 'B', 'C', 'D', 'E', 'F', 'G', 'H', 'I', 'J', 'K', 'M']
            self.ListaPCP['A'] = self.ListaPCP['A'].fillna('-SEM CODIGO-')
            self.ListaPCP['B'] = self.ListaPCP['B'].fillna('-SEM DESCRICAO-')
            self.ListaPCP['C'] = self.ListaPCP['C'].fillna('-SEM REFERENCIA-')
            self.ListaPCP['D'] = self.ListaPCP['D'].fillna('-PESO NAO PREENCHIDO-')
            self.ListaPCP['E'] = self.ListaPCP['E'].fillna('-QTDE NAO PREENCHIDA-')
            self.ListaPCP['F'] = self.ListaPCP['F'].fillna('-MATERIAL NAO DEFINIDO-')
            self.ListaPCP['G'] = self.ListaPCP['G'].fillna('PROCESSO NAO DEFINIDO')
            self.ListaPCP['H'] = self.ListaPCP['H'].fillna('SEM DIMENSIONAMENTO')
            self.ListaPCP['I'] = self.ListaPCP['I'].fillna('-TRATAMENTO VAZIO-')
            self.ListaPCP['J'] = self.ListaPCP['J'].fillna('-TRATAMENTO VAZIO-')
            self.ListaPCP['K'] = self.ListaPCP['K'].fillna('USINAGEM NAO DEFINIDA')
            self.ListaPCP['M'] = self.ListaPCP['M'].fillna('FORNECEDOR NAO DEFINIDO')

            # Cria as listas com os arquivos de desenhos existentes
            self.Desenhos_Path = (self.lineEdit_Desenhos.text())
            # self.ListaArquivos = os.listdir(self.Desenhos_Path)
            for dirpath, subdirs, files in os.walk(self.Desenhos_Path):
                for file in files:
                    self.ListaArquivos.append(os.path.normpath(os.path.join(dirpath, file)))

        except Exception as e:
            traceback_str = traceback.format_exc()
            stack = traceback.extract_stack()
            function_name = stack[-2].name
            print(f"Ocorreu um erro na função '{function_name}' e o código foi abortado.")
            print("Erro: " + repr(e))
            print("Descrição do erro: " + str(e))
            print("Traceback: " + traceback_str)

    # Função que aciona as funções de criação de pastas
    def gerar_pastas(self):

        # Previne a execucao da funcao se nao for selecionada uma opcao de codigo
        if not (self.radioButton_codigo.isChecked() or self.radioButton_referencia.isChecked()):
            print("Informe qual é a coluna com o nome dos arquivos: Código ou Referência.")
            return

        # Carrega os arquivos basicos
        self.ler_arquivo_pcp()

        # Cria as pastas dos processos completas caso a opcao seja selecionada
        if self.checkBox_Processos.isChecked():
            self.criar_pastas_processos()
        else:
            print("Pastas de PROCESSOS não foram criadas, pois não foi solicitado.")

        # Cria a pasta de pintura completa caso a opcao seja selecionada
        if self.checkBox_Pintura.isChecked():
            self.criar_pasta_pintura()
        else:
            print("Pasta de PINTURA não foi criada, pois não foi solicitado.")

        # Cria a pasta de usinagem completa caso a opcao seja selecionada
        if self.checkBox_Usinagem.isChecked():
            self.criar_pasta_usinagem()
        else:
            print("Pasta de USINAGEM não foi criada, pois não foi solicitado.")

        # Cria as pastas dos fornecedores de usinagem caso a opcao seja selecionada
        if self.checkBox_FornUsi.isChecked():
            self.criar_pastas_fornusi()
        else:
            print("Pastas de FORN. USINAGEM não foram criadas, pois não foi solicitado.")

    # Função que cria as pastas de processo de fabricação básicas
    def criar_pastas_processos(self):
        try:
            # Cria a lista de processos sem repetição, sem desmembrar processos compostos
            lista_processos = list(set(self.ListaPCP['G'].dropna().tolist()))

            # Separa os processos compostos e cria a lista final
            lista_processos_temporaria = []  # Lista vazia para armazenar os itens separados
            for item in lista_processos:
                if '+' in item:
                    substrings = item.split("+")
                    lista_processos_temporaria.extend(substrings)
                else:
                    lista_processos_temporaria.append(item)
            lista_processos = list(set(lista_processos_temporaria))

            # Mostra resultado da lista final
            print_processos = ', '.join(str(x) for x in lista_processos)
            print("Lista de processos:", print_processos)

            if self.radioButton_codigo.isChecked():
                planilha_pcp = self.ListaPCP.drop(labels=['B', 'C', 'D', 'E', 'F', 'H', 'I', 'J', 'K'], axis=1)
            elif self.radioButton_referencia.isChecked():
                planilha_pcp = self.ListaPCP.drop(labels=['A', 'B', 'D', 'E', 'F', 'H', 'I', 'J', 'K'], axis=1)
            else:
                print("Coluna com os códigos dos desenhos não foi selecionada. Favor selecionar uma opção válida.")
                return

            # Itera por cada processo, criando a pasta e copiando os desenhos correpondentes
            for processo in lista_processos:

                pasta_processo = os.path.join((os.path.dirname(self.PCP_Path)), processo)  # Nome da pasta a ser criada
                # Confere se a pasta do processo já existe antes de criá-la.
                # Caso ela já exista, o programa informa o usuário e pula para a próxima iteração
                if os.path.exists(pasta_processo):
                    print(f"A pasta '{processo}' já existe, ela não será criada novamente e este processo será pulado.")
                    continue
                else:
                    os.mkdir(pasta_processo)  # Cria a pasta do processo

                # Chama a funcao para criar a planilha de quantidades
                self.gerar_planilhas(processo, pasta_processo)

                # Filtra a planilha pelo processo correspondente da iteracao
                condicao_processo = lambda z: True if processo in z else False
                codigos_por_processo = planilha_pcp.loc[planilha_pcp['G'].str.split('+').apply(condicao_processo)]

                # Orienta o programa a buscar os codigos conforme a opcao do usuario
                if self.radioButton_codigo.isChecked():
                    codigos = list(codigos_por_processo['A'])
                elif self.radioButton_referencia.isChecked():
                    codigos = list(codigos_por_processo['C'])
                else:
                    print("Coluna com os códigos dos desenhos não foi selecionada. Favor selecionar uma opção válida.")
                    return

                # Procura os arquivos correspondentes a cada código e copia para a pasta
                self.copiar_arquivos(codigos, pasta_processo)

                # Chama a funcao para compactar os arquivos.
                self.criar_arquivo_compactado(processo, pasta_processo)

            print("Pastas de PROCESSOS criadas com sucesso.")

        except Exception as e:
            traceback_str = traceback.format_exc()
            stack = traceback.extract_stack()
            function_name = stack[-2].name
            print(f"Ocorreu um erro na função '{function_name}' e o código foi abortado.")
            print("Erro: " + repr(e))
            print("Descrição do erro: " + str(e))
            print("Traceback: " + traceback_str)

    # Função que cria a pasta de PINTURA
    def criar_pasta_pintura(self):
        try:
            # Filtra a planilha para conter apenas as linhas com itens de pintura
            listapintura = self.ListaPCP.loc[self.ListaPCP['J'].str.contains("fp|ral|pint", case=False, na=False)]

            # Cria lista com os códigos das peças de pintura, dependendo da seleção do usuário
            if self.radioButton_codigo.isChecked():
                listapintura = list(set(listapintura['A'].tolist()))
            elif self.radioButton_referencia.isChecked():
                listapintura = list(set(listapintura['C'].tolist()))

            # Caminho para pasta de pintura
            pasta_pintura = os.path.join((os.path.dirname(self.PCP_Path)), "PINTURA")

            # Verifica se a pasta já existe
            if os.path.exists(pasta_pintura):
                print('A pasta "PINTURA" já existe, o procedimento será abortado.')
                return
            else:
                os.mkdir(pasta_pintura)  # Cria a pasta de pintura

            processo = 'Pintura'

            # Chama a funcao para criar a planilha de quantidades
            self.gerar_planilhas(processo, pasta_pintura)

            # Copia os arquivos encontrados
            self.copiar_arquivos(listapintura, pasta_pintura)

            # Chama a funcao para criar o arquivo compactado
            self.criar_arquivo_compactado(processo, pasta_pintura)

            print('Pasta de PINTURA criada com sucesso.')

        except Exception as e:
            traceback_str = traceback.format_exc()
            stack = traceback.extract_stack()
            function_name = stack[-2].name
            print(f"Ocorreu um erro na função '{function_name}' e o código foi abortado.")
            print("Erro: " + repr(e))
            print("Descrição do erro: " + str(e))
            print("Traceback: " + traceback_str)

    # Função que cria a pasta de USINAGEM
    def criar_pasta_usinagem(self):
        try:
            # Cria a lista de processos sem itens repetidos e sem desmembrar processos compostos
            lista_usinagem = set(self.ListaPCP['K'].dropna().tolist())
            lista_usinagem.discard("-")
            lista_usinagem = list(lista_usinagem)

            # Caminho da pasta de Usinagem
            pasta_usinagem = os.path.join((os.path.dirname(self.PCP_Path)), "USINAGEM")

            # Verifica se a pasta já existe
            if os.path.exists(pasta_usinagem):
                print('A pasta "USINAGEM" já existe, o procedimento será abortado.')
                return
            else:
                os.mkdir(pasta_usinagem)  # Cria a pasta de usinagem

            # Filtra a lista com os códigos das peças dependendo da seleção do usuário
            if self.radioButton_codigo.isChecked():
                planilha = self.ListaPCP.drop(labels=['B', 'C', 'D', 'E', 'F', 'G', 'H', 'I', 'J', 'M'], axis=1)
            elif self.radioButton_referencia.isChecked():
                planilha = self.ListaPCP.drop(labels=['A', 'B', 'D', 'E', 'F', 'G', 'H', 'I', 'J', 'M'], axis=1)
            else:
                print("Coluna com os códigos dos desenhos não foi selecionada. Favor selecionar uma opção válida.")
                return

            for porte_usi in lista_usinagem:
                # Caminho completo da pasta a ser criada
                pasta_porte_usinagem = os.path.join(pasta_usinagem, porte_usi)
                # Verifica se a pasta já existe
                if os.path.exists(pasta_porte_usinagem):
                    print(f"A pasta '{porte_usi}' já existe, ela não será criada novamente.")
                    print("O processo de USINAGEM será abortado.")
                    return
                else:
                    os.mkdir(pasta_porte_usinagem)  # Cria a pasta do processo

                # Filtra a planilha pelo processo correspondente da iteracao
                codigos_usi = planilha.loc[planilha['K'].str.contains(porte_usi, case=False, na=False, regex=False)]

                # Cria a lista com os códigos, filtrados pelo processo correspondente da iteracao
                if self.radioButton_codigo.isChecked():
                    codigos = list(codigos_usi['A'])
                elif self.radioButton_referencia.isChecked():
                    codigos = list(codigos_usi['C'])
                else:
                    print("Coluna com os códigos dos desenhos não foi selecionada. Favor selecionar uma opção válida.")
                    return

                # Copia os arquivos encontrados
                self.copiar_arquivos(codigos, pasta_porte_usinagem)

            # Cria pasta contendo todas as usinagens de pequeno porte
            # Filtra a planilha
            pre_lista_usi_p = planilha.loc[~planilha['K'].str.contains("portal|-", case=False, na=False)]

            if self.radioButton_codigo.isChecked():
                lista_usi_p = list(pre_lista_usi_p['A'])
            elif self.radioButton_referencia.isChecked():
                lista_usi_p = list(pre_lista_usi_p['C'])
            else:
                print("Coluna com os códigos dos desenhos não foi selecionada. Favor selecionar uma opção válida.")
                return

            pasta_usi_p = os.path.join(pasta_usinagem, "Usinagem_P")
            os.mkdir(pasta_usi_p)
            self.copiar_arquivos(lista_usi_p, pasta_usi_p)
            self.criar_arquivo_compactado("Usinagem_P", pasta_usi_p)
            self.gerar_planilhas('Usinagem_P', pasta_usi_p)

            print("Pasta de USINAGEM criada com sucesso.")

        except Exception as e:
            traceback_str = traceback.format_exc()
            stack = traceback.extract_stack()
            function_name = stack[-2].name
            print(f"Ocorreu um erro na função '{function_name}' e o código foi abortado.")
            print("Erro: " + repr(e))
            print("Descrição do erro: " + str(e))
            print("Traceback: " + traceback_str)

    def criar_pastas_fornusi(self):
        try:
            # Cria a lista de fornecedores
            lista_fornusi = set(self.ListaPCP['M'].dropna().tolist())
            lista_fornusi.discard("-")
            lista_fornusi = list(lista_fornusi)

            # Filtra a lista com os códigos das peças dependendo da seleção do usuário
            if self.radioButton_codigo.isChecked():
                planilha = self.ListaPCP.drop(labels=['B', 'C', 'D', 'E', 'F', 'G', 'H', 'I', 'J', 'K'], axis=1)
            elif self.radioButton_referencia.isChecked():
                planilha = self.ListaPCP.drop(labels=['A', 'B', 'D', 'E', 'F', 'G', 'H', 'I', 'J', 'K'], axis=1)
            else:
                print("Coluna com os códigos dos desenhos não foi selecionada. Favor selecionar uma opção válida.")
                return

            for fornecedor in lista_fornusi:
                # Criacao da pasta
                pasta_usinagem = os.path.join((os.path.dirname(self.PCP_Path)), "USINAGEM")
                if os.path.exists(pasta_usinagem):
                    pasta_fornusi = os.path.join(pasta_usinagem, fornecedor)
                    if os.path.exists(pasta_fornusi):
                        print(f"A pasta '{fornecedor}' já existe, o procedimento será abortado.")
                        return
                    else:
                        os.mkdir(pasta_fornusi)
                else:
                    os.mkdir(pasta_usinagem)  # Cria a pasta de usinagem
                    pasta_fornusi = os.path.join(pasta_usinagem, fornecedor)
                    os.mkdir(pasta_fornusi)

                # Filtra a planilha pelo fornecedor correspondente da iteracao
                codigos_fornusi = planilha.loc[planilha['M'].str.contains(fornecedor, case=False, na=False, regex=False)]

                # Cria a lista com os códigos, filtrados pelo processo correspondente da iteracao
                if self.radioButton_codigo.isChecked():
                    codigos = list(codigos_fornusi['A'])
                elif self.radioButton_referencia.isChecked():
                    codigos = list(codigos_fornusi['C'])
                else:
                    print("Coluna com os códigos dos desenhos não foi selecionada. Favor selecionar uma opção válida.")
                    return

                # Copia os arquivos encontrados
                self.copiar_arquivos(codigos, pasta_fornusi)
                self.gerar_planilhas(fornecedor, pasta_fornusi, True)
                self.criar_arquivo_compactado(fornecedor, pasta_fornusi)

            print('Pastas de FORN. USINAGEM criadas com sucesso.')

        except Exception as e:
            traceback_str = traceback.format_exc()
            stack = traceback.extract_stack()
            function_name = stack[-2].name
            print(f"Ocorreu um erro na função '{function_name}' e o código foi abortado.")
            print("Erro: " + repr(e))
            print("Descrição do erro: " + str(e))
            print("Traceback: " + traceback_str)

    # Função que verifica se existem arquivos faltando na pasta de desenhos
    def verificar_arquivos_desenhos_pcp(self):
        try:
            # Le o arquivo PCP para ter os codigos a serem verificados
            self.ler_arquivo_pcp()

            if self.radioButton_codigo.isChecked():
                lista_codigos = list(set(self.ListaPCP['A'].dropna().tolist()))
            elif self.radioButton_referencia.isChecked():
                lista_codigos = list(set(self.ListaPCP['C'].dropna().tolist()))
            else:
                print("Coluna com os códigos dos desenhos não foi selecionada. Favor selecionar uma opção válida.")
                return

            # Variavel com os formatos de arquivos utilizados
            tipos_arquivos = [".igs", ".pdf", ".dwg"]

            # Caminho do relatorio de arquivos faltantes, para checagem.
            caminho_log = os.path.join(os.path.dirname(self.PCP_Path), "log_arquivos_faltantes.txt")

            # Gerar mensagem inicial em formato de email
            mensagem_inicial = "Prezado(a) Projetista,\n\nNão conseguimos encontrar na pasta informada alguns " \
                               "arquivos de desenhos das peças dessa lista. Pedimos a gentileza de providenciar " \
                               "os arquivos na pasta para que possamos prosseguir com o processo de compra.\nA seguir" \
                               " está a lista dos arquivos faltantes.\n\n"
            flag_faltas = False

            for codigo in lista_codigos:
                codigo = str(codigo)
                arquivos = [arquivo.lower() for arquivo in self.ListaArquivos if re.search(codigo, arquivo, re.I)]
                # Verificar extensões em arquivo_encontrado
                extensoes_encontradas = [os.path.splitext(arquivo)[1] for arquivo in arquivos]
                # Verificar extensões ausentes
                extensoes_ausentes = [extensao for extensao in tipos_arquivos if
                                      extensao not in extensoes_encontradas]
                # Gerar mensagem das extensões ausentes
                if extensoes_ausentes:
                    flag_faltas = True
                    if os.path.exists(caminho_log):
                        with open(caminho_log, "a") as arquivo_log:
                            for extensao_ausente in extensoes_ausentes:
                                mensagem = f"Arquivo '{extensao_ausente}' não encontrado para o código '{codigo}'"
                                print(mensagem)
                                arquivo_log.write(mensagem + "\n")
                    else:
                        with open(caminho_log, mode='a') as arquivo_log:
                            arquivo_log.write(mensagem_inicial)
                            for extensao_ausente in extensoes_ausentes:
                                mensagem = f"Arquivo '{extensao_ausente}' não encontrado para o código '{codigo}'"
                                print(mensagem)
                                arquivo_log.write(mensagem + "\n")

            if flag_faltas:
                print("Salvo arquivo 'log_arquivos_faltantes.txt' contendo mensagem pronta"
                      " para informar o projetista responsável.")
            else:
                print('Não foi encontrado nenhum arquivo faltante, com base nos critérios fornecidos.')

        except Exception as e:
            traceback_str = traceback.format_exc()
            stack = traceback.extract_stack()
            function_name = stack[-2].name
            print(f"Ocorreu um erro na função '{function_name}' e o código foi abortado.")
            print("Erro: " + repr(e))
            print("Descrição do erro: " + str(e))
            print("Traceback: " + traceback_str)

    def ler_arquivo_csv(self):
        try:
            local_csv = (self.lineEdit_ArquivoPMS.text())
            self.Arquivo_CSV = pd.read_csv(
                filepath_or_buffer=local_csv,
                sep=';',
                header=0,
                usecols=[1, 2, 3, 6, 7, 8, 11, 12, 13, 14],
                dtype=str,
                on_bad_lines='warn',
                encoding='ISO-8859-1'
            )
            print('Arquivo CSV carregado.')

            # Renomeia as colunas do dataframe
            self.Arquivo_CSV.columns = ['B', 'C', 'D', 'G', 'H', 'I', 'L', 'M', 'N', 'O']

        except Exception as e:
            traceback_str = traceback.format_exc()
            stack = traceback.extract_stack()
            function_name = stack[-2].name
            print(f"Ocorreu um erro na função '{function_name}' e o código foi abortado.")
            print("Erro: " + repr(e))
            print("Descrição do erro: " + str(e))
            print("Traceback: " + traceback_str)

    def listar_arquivos_pasta(self):
        try:
            local_arquivos_desenhos = (self.lineEdit_Desenhos.text())
            self.ListaArquivos = os.listdir(local_arquivos_desenhos)
            print('Diretorio de arquivos carregado.')

        except Exception as e:
            traceback_str = traceback.format_exc()
            stack = traceback.extract_stack()
            function_name = stack[-2].name
            print(f"Ocorreu um erro na função '{function_name}' e o código foi abortado.")
            print("Erro: " + repr(e))
            print("Descrição do erro: " + str(e))
            print("Traceback: " + traceback_str)

    # Função que verifica se existem arquivos faltando na pasta de desenhos
    def verificar_arquivos_desenhos_pms(self):
        try:
            # Executa funcao que le o arquivo .csv
            self.ler_arquivo_csv()

            # Executa funcao de leitura dos arquivos no diretorio
            self.listar_arquivos_pasta()

            # Filtra a planilha inteira deixando apenas os itens de fabricação (BR-PROD-MA)
            arquivo_csv = self.Arquivo_CSV.loc[self.Arquivo_CSV['I'].str.contains("PROD", case=False, na=False)]\
                .drop(labels=['I'], axis=1).reset_index(drop=True)

            # Cria lista de códigos com base na escolha do usuario
            if self.radioButton_codigo.isChecked():
                lista_codigos = list(set(arquivo_csv['B'].dropna().tolist()))
            elif self.radioButton_referencia.isChecked():
                lista_codigos = list(set(arquivo_csv['D'].dropna().tolist()))
            else:
                print("Coluna com os códigos dos desenhos não foi selecionada. Favor selecionar uma opção válida.")
                return

            # Variavel com os formatos de arquivos utilizados
            tipos_arquivos = [".igs", ".pdf", ".dwg"]

            # Caminho do relatorio de arquivos faltantes, para checagem.
            caminho_log = os.path.join(os.path.dirname(self.lineEdit_ArquivoPMS.text()), "log_arquivos_faltantes.txt")

            # Gerar mensagem inicial em formato de email
            mensagem_inicial = "Prezado(a) Projetista,\n\nNão conseguimos encontrar na pasta informada alguns " \
                               "arquivos de desenhos das peças dessa lista. Pedimos a gentileza de providenciar " \
                               "os arquivos na pasta para que possamos prosseguir com o processo de compra.\nA seguir" \
                               " está a lista dos arquivos faltantes.\n\n"
            flag_faltas = False

            for codigo in lista_codigos:
                codigo = str(codigo)
                arquivos = [arquivo.lower() for arquivo in self.ListaArquivos if re.search(codigo, arquivo, re.I)]
                # Verificar extensões em arquivo_encontrado
                extensoes_encontradas = [os.path.splitext(arquivo)[1] for arquivo in arquivos]
                # Verificar extensões ausentes
                extensoes_ausentes = [extensao for extensao in tipos_arquivos if
                                      extensao not in extensoes_encontradas]
                # Gerar mensagem das extensões ausentes
                if extensoes_ausentes:
                    flag_faltas = True
                    if os.path.exists(caminho_log):
                        with open(caminho_log, "a") as arquivo_log:
                            for extensao_ausente in extensoes_ausentes:
                                mensagem = f"Arquivo '{extensao_ausente}' não encontrado para o código '{codigo}'"
                                print(mensagem)
                                arquivo_log.write(mensagem + "\n")
                    else:
                        with open(caminho_log, mode='a') as arquivo_log:
                            arquivo_log.write(mensagem_inicial)
                            for extensao_ausente in extensoes_ausentes:
                                mensagem = f"Arquivo '{extensao_ausente}' não encontrado para o código '{codigo}'"
                                print(mensagem)
                                arquivo_log.write(mensagem + "\n")

            if flag_faltas:
                print("Salvo arquivo 'log_arquivos_faltantes.txt' contendo mensagem pronta"
                      " para informar o projetista responsável.")
            else:
                print('Não foi encontrado nenhum arquivo faltante, com base nos critérios fornecidos.')

        except Exception as e:
            traceback_str = traceback.format_exc()
            stack = traceback.extract_stack()
            function_name = stack[-2].name
            print(f"Ocorreu um erro na função '{function_name}' e o código foi abortado.")
            print("Erro: " + repr(e))
            print("Descrição do erro: " + str(e))
            print("Traceback: " + traceback_str)

    # FUNCAO NAO IMPLEMENTADA CORRETAMENTE!!! CODIGO PROVISORIO!
    # Informacao para revisao futura:
    # Objetivo para a funcao: verificar os desenhos que sofreram alteracao, informar o usuario e atualizar as pastas.
    def atualizar(self):
        self.limpeza()
        self.gerar_pastas()

    # Função que deleta todas as pastas criadas na pasta do PCP.
    def limpeza(self):
        try:
            diretorio_pcp = os.path.dirname(self.lineEdit_ArquivoPCP.text())
            if not any(entrada.is_dir() for entrada in os.scandir(diretorio_pcp)):
                print("Nao existem pastas para excluir!")
                return
            for entrada in os.scandir(diretorio_pcp):
                if entrada.is_dir() and entrada.path != diretorio_pcp:
                    shutil.rmtree(entrada.path, ignore_errors=True)
            print("Limpeza concluida.")
        except Exception as e:
            traceback_str = traceback.format_exc()
            stack = traceback.extract_stack()
            function_name = stack[-2].name
            print(f"Ocorreu um erro na função '{function_name}' e o código foi abortado.")
            print("Erro: " + repr(e))
            print("Descrição do erro: " + str(e))
            print("Traceback: " + traceback_str)

    def copiar_arquivos(self, lista_codigos, pasta_destino):
        try:
            for codigo in lista_codigos:
                # Cria lista de arquivos que contem o codigo fornecido
                lista_arquivos = [arquivo for arquivo in self.ListaArquivos if re.search(codigo, arquivo, re.I)]
                if lista_arquivos:
                    for arquivo in lista_arquivos:
                        # Cria o nome do arquivo futuro
                        nomearquivo = os.path.basename(arquivo)
                        arquivo_final = os.path.join(pasta_destino, nomearquivo)
                        # Confere se o desenho a ser copiado ja existe
                        if os.path.exists(arquivo_final):
                            print(f'O arquivo "{nomearquivo}" já existe na pasta "{pasta_destino}" e não será copiado.')
                            continue
                        else:
                            shutil.copy(arquivo, pasta_destino)
                else:
                    print(f'Nenhum arquivo encontrado para o código "{codigo}" para o processo "{os.path.basename(pasta_destino)}".')
                    continue
                    
        except Exception as e:
            traceback_str = traceback.format_exc()
            stack = traceback.extract_stack()
            function_name = stack[-2].name
            print(f"Ocorreu um erro na função '{function_name}' e o código foi abortado.")
            print("Erro: " + repr(e))
            print("Descrição do erro: " + str(e))
            print("Traceback: " + traceback_str)

    def gerar_planilhas(self, processo, pasta_destino, flag_usi=False):
        try:
            # Carrega o arquivo '.xlsm' do PCP prenchido
            excel_pcp = load_workbook(self.PCP_Path, keep_vba=False)
            manter_planilha = 'Principal'
            for nome_planilha in excel_pcp.sheetnames:
                if nome_planilha not in manter_planilha:
                    del excel_pcp[nome_planilha]
            excel_pcp['Principal'].title = "QUANTITATIVO"

            # Variavel lista para armazenar as linhas que vao ser preservadas.
            linhas_filtradas = []

            # Condicao "if" para verificar se a planilha vai ser para pasta de pintura.
            # Isso acontece pois a variavel processo nao se aplica para a pasta de pintura.
            if 'pint' in processo.lower():
                termos = ['pint', 'ral', 'fp']
                for index, linha in enumerate(excel_pcp['QUANTITATIVO'].iter_rows(min_row=2, values_only=True), start=2):
                    if not any(termo in str(linha[9]).lower() for termo in termos):
                        linhas_filtradas.append(index)
            elif 'usinagem_p' in processo.lower():
                termos = ['portal', '-']
                for index, linha in enumerate(excel_pcp['QUANTITATIVO'].iter_rows(min_row=2, values_only=True), start=2):
                    if any(termo in str(linha[10]).lower() for termo in termos):
                        linhas_filtradas.append(index)
            elif flag_usi == True:
                for index, linha in enumerate(excel_pcp['QUANTITATIVO'].iter_rows(min_row=2, values_only=True), start=2):
                    if processo not in str(linha[12]):
                        linhas_filtradas.append(index)
            else:
                for index, linha in enumerate(excel_pcp['QUANTITATIVO'].iter_rows(min_row=2, values_only=True), start=2):
                    if processo not in str(linha[6]).split("+"):
                        linhas_filtradas.append(index)

            # Exclui da planilha todas as linhas que não contém o processo.
            for index_linhas in reversed(linhas_filtradas):
                excel_pcp['QUANTITATIVO'].delete_rows(index_linhas)

            # Remove as cores da planilha exceto na primeira linha
            no_fill = PatternFill(start_color="FFFFFF", end_color="FFFFFF", fill_type="solid")
            plan = excel_pcp['QUANTITATIVO']
            for row in plan.iter_rows(min_row=2, max_row=plan.max_row, min_col=1, max_col=plan.max_column):
                for cell in row:
                    cell.fill = no_fill

            # Deleta as demais colunas e remove o autofiltro do cabecalho.
            excel_pcp['QUANTITATIVO'].delete_cols(11, excel_pcp['QUANTITATIVO'].max_column)
            if 'pint' in processo.lower():
                excel_pcp['QUANTITATIVO'].delete_cols(6, 4)
            excel_pcp['QUANTITATIVO'].auto_filter.ref = None

            # Cria o nome do arquivo com base no processo.
            if '-PCP.xlsm' in os.path.basename(self.PCP_Path):
                nome_arquivo = os.path.basename(self.PCP_Path).replace('-PCP.xlsm', str('_' + processo + '.xlsx'))
            else:
                nome_arquivo = str(processo + '.xlsx')

            # Cria o caminho completo do arquivo a ser salvo.
            arquivo_quantitativo = os.path.join(pasta_destino, nome_arquivo)

            # Salva e fecha o arquivo.
            excel_pcp.save(arquivo_quantitativo)
            excel_pcp.close()

        except Exception as e:
            traceback_str = traceback.format_exc()
            stack = traceback.extract_stack()
            function_name = stack[-2].name
            print(f"Ocorreu um erro na função '{function_name}' e o código foi abortado.")
            print("Erro: " + repr(e))
            print("Descrição do erro: " + str(e))
            print("Traceback: " + traceback_str)

    def get_total_file_size(self, file_list):
        total_size = 0
        for file_path in file_list:
            total_size += os.path.getsize(file_path)
        return total_size

    def criar_arquivo_compactado(self, nome_processo, pasta_origem):
        try:
            # Define the maximum allowed size in bytes (50MB)
            max_size_bytes = 50 * 1024 * 1024  # 20MB in bytes

            # Obtem a lista de arquivos no diretorio de origem
            arq_comp = os.listdir(pasta_origem)

            # Filtra e remove arquivos com a extensão ".xlsm"
            arq_comp = [arquivo for arquivo in arq_comp if not arquivo.endswith(".xlsx")]

            # Calculate the total size of files to be compressed
            total_size = self.get_total_file_size([os.path.join(pasta_origem, arquivo) for arquivo in arq_comp])

            # Check if the total size is below the limit
            if total_size <= max_size_bytes:
                # Cria o nome do arquivo com base no processo.
                if '-PCP.xlsm' in os.path.basename(self.PCP_Path):
                    nome_arquivo = os.path.basename(self.PCP_Path).replace('-PCP.xlsm', str('_' + nome_processo + '.7z'))
                else:
                    nome_arquivo = str(nome_processo + '.7z')

                # Cria o caminho completo do arquivo comprimido
                arquivo_7z = os.path.join(pasta_origem, nome_arquivo)

                # Cria o comando para compactar os arquivos em um arquivo ".7z" com compactacao Ultra
                comando = ['7z', 'a', '-t7z', '-mx', arquivo_7z] + [os.path.join(pasta_origem, arquivo) for arquivo in arq_comp]

                # Executa o comando usando o "subprocess"
                subprocess.run(comando, check=True)
            else:
                print("Tamanho total dos arquivos para compactação excedeu o limite. Arquivo compactado não foi criado.")

        except Exception as e:
            traceback_str = traceback.format_exc()
            stack = traceback.extract_stack()
            function_name = stack[-2].name
            print(f"Ocorreu um erro na função '{function_name}' e o código foi abortado.")
            print("Erro: " + repr(e))
            print("Descrição do erro: " + str(e))
            print("Traceback: " + traceback_str)

    def gerar_arquivo_pcp_csv(self):
        try:
            # COLUNAS DO CSV
            # 1-B = Código
            # 2-C = Descrição
            # 3-D = Referencia
            # 6-G = Peso
            # 7-H = Quantidade
            # 8-I = Tipo (BR-PROD-MA)
            # 9-L = Status
            # 12-M = Material
            # 13-N = Tratamento térmico
            # 14-O = Tratamento superficial

            proj_len = len(self.lineEdit_NomeProjeto.text()) > 0
            list_len = len(self.lineEdit_NomeLista.text()) > 0

            if not proj_len and list_len:
                print("Para prosseguir preencha projeto e lista corretamente.")
                return

            # Caminhos para os arquivos básicos, fornecidos pelo operador
            pmspath = (self.lineEdit_ArquivoPMS.text())
            caminho_pcp_padrao = (self.lineEdit_ArquivoPCPPadrao.text())

            # Cria uma cópia do arquivo PCP em branco para a pasta
            pasta_projeto = os.path.dirname(os.path.dirname(os.path.dirname(os.path.dirname(pmspath))))
            pasta_fabricacao = os.path.join(pasta_projeto, "23_PCP_FABRICACAO")
            projeto = (self.lineEdit_NomeProjeto.text()).zfill(6)
            lista = (self.lineEdit_NomeLista.text()).zfill(8)
            nome_pasta_pcp = f'{projeto}({lista})'
            pasta_pcp = os.path.join(pasta_fabricacao, nome_pasta_pcp)
            if os.path.exists(pasta_pcp):
                print(f"A pasta {nome_pasta_pcp} já existe, procedimento abortado.")
                return
            else:
                os.mkdir(pasta_pcp)
            nome_arquivo_pcp = f'{projeto}({lista})-PCP.xlsm'
            arquivo_pcp_zero = os.path.join(pasta_pcp, nome_arquivo_pcp)
            shutil.copy(caminho_pcp_padrao, arquivo_pcp_zero)
            print(f'Arquivo PCP padrão copiado para a pasta {pasta_pcp}')

            # Comando de leitura do arquivo CSV
            self.ler_arquivo_csv()
            arquivo_csv = self.Arquivo_CSV
            print('Arquivo CSV carregado.')

            # Dicionario para colar as colunas nos locais corretos da planilha
            dict_colunas = {
                'B': 1,
                'C': 2,
                'D': 3,
                'G': 4,
                'H': 5,
                'I': 8,
                'M': 6,
                'N': 9,
                'O': 10
            }

            # Carrega o arquivo XLSM
            arquivo_excel = load_workbook(arquivo_pcp_zero, keep_vba=True)
            # Seleciona a planilha onde serão adicionados os dados do CSV
            planilha_principal = arquivo_excel['Principal']
            planilha_resumo = arquivo_excel['Resumo']

            ###############################################################################
            # SEQUENCIA DE COMANDOS PARA FORMATAR OS DADOS ANTES DE COLAR NA PLANILHA PCP #
            ###############################################################################

            # Renomeia as colunas do dataframe
            arquivo_csv.columns = ['B', 'C', 'D', 'G', 'H', 'I', 'L', 'M', 'N', 'O']

            # Transforma a coluna de peso em número decimal
            arquivo_csv['G'] = pd.to_numeric(arquivo_csv['G'], errors='coerce').astype(float).round(3)

            # Transforma a coluna de quantidade em números inteiros
            arquivo_csv['H'] = pd.to_numeric(arquivo_csv['H'], errors='coerce').fillna(0).astype(int)

            # Altera a coluna de descrição para apenas a primeira letra de cada linha ser maiúscula
            arquivo_csv['C'] = arquivo_csv['C'].str.capitalize()

            # Altera a coluna de descrição removendo os espaços em branco no início e final de cada nome
            arquivo_csv = arquivo_csv.applymap(lambda x: x.strip() if isinstance(x, str) else x)

            # Altera a coluna de descrição removendo acentuação
            arquivo_csv = arquivo_csv.applymap(lambda x: unidecode(x) if isinstance(x, str) else x)

            # Filtra a planilha inteira deixando apenas os itens com Status:Production
            arquivo_csv = arquivo_csv.loc[arquivo_csv['L'].str.contains("production", case=False, na=False)] \
                .drop(labels=['L'], axis=1).reset_index(drop=True)

            keywords = ['PROD', 'WELD']

            # Filtra novamente a planilha inteira deixando apenas os itens de fabricação (BR-PROD-MA)
            arquivo_csv = arquivo_csv.loc[arquivo_csv['I'].str.contains('|'.join(keywords), case=False, na=False)]\
                .drop(labels=['I'], axis=1).reset_index(drop=True)

            # Altera as colunas de 'codigos' e 'referencias' para 'int64' ou 'string'
            arquivo_csv['B'] = arquivo_csv['B'].astype('int64') if arquivo_csv['B']\
                .apply(lambda x: str(x).isdigit()).all() else arquivo_csv['B'].astype(str)
            arquivo_csv['D'] = arquivo_csv['D'].astype('int64') if arquivo_csv['D'] \
                .apply(lambda x: str(x).isdigit()).all() else arquivo_csv['D'].astype(str)

            print('Informacoes do PMS formatadas.')

            ###############################################################################
            #                  FIM DA FORMATACAO DOS DADOS DO DATAFRAME                   #
            ###############################################################################

            # Variável para armazenar a última linha preenchida
            ultima_linha_preenchida = 0

            # Preenche as celulas recursivamente com as informações do dataframe 'CSV'
            for index_linha, linha in arquivo_csv.iterrows():
                for index_coluna, valor_celula in linha.items():
                    novo_index_coluna = dict_colunas[index_coluna]
                    planilha_principal.cell(row=index_linha + 2, column=novo_index_coluna, value=valor_celula)
                ultima_linha_preenchida = index_linha + 2

            print('Arquivo excel do PCP preenchido.')

            # Deleta as linhas que sobram na planilha
            ultima_linha = planilha_principal.max_row - 1
            for linha in range((ultima_linha - 1), ultima_linha_preenchida, -1):
                planilha_principal.delete_rows(linha)

            print('Deletadas as linhas extras.')

            # Corrige as fórmulas de soma com os totais após deletar as linhas
            ultima_linha = planilha_principal.max_row - 1
            planilha_principal.cell(row=ultima_linha, column=14).value = f'=SUM(N2:N{ultima_linha - 1})'
            planilha_principal.cell(row=ultima_linha, column=15).value = f'=SUM(O2:O{ultima_linha - 1})'
            planilha_principal.cell(row=ultima_linha, column=16).value = f'=SUM(P2:P{ultima_linha - 1})'
            planilha_principal.cell(row=ultima_linha, column=17).value = f'=SUM(Q2:Q{ultima_linha - 1})'
            planilha_principal.cell(row=ultima_linha, column=18).value = f'=SUM(R2:R{ultima_linha - 1})'
            planilha_principal.cell(row=ultima_linha, column=19).value = f'=SUM(S2:S{ultima_linha - 1})'
            planilha_principal.cell(row=ultima_linha, column=20).value = f'=SUM(T2:T{ultima_linha - 1})'
            planilha_principal.cell(row=ultima_linha, column=21).value = f'=SUM(U2:U{ultima_linha - 1})'
            planilha_principal.cell(row=ultima_linha, column=22).value = f'=SUM(V2:V{ultima_linha - 1})'
            planilha_principal.cell(row=ultima_linha, column=23).value = f'=SUM(W2:W{ultima_linha - 1})'
            planilha_principal.cell(row=ultima_linha, column=24).value = f'=SUM(X2:X{ultima_linha - 1})'
            planilha_principal.cell(row=ultima_linha, column=25).value = f'=SUM(Y2:Y{ultima_linha - 1})'
            planilha_principal.cell(row=ultima_linha, column=26).value = f'=SUM(Z2:Z{ultima_linha - 1})'
            planilha_principal.cell(row=ultima_linha, column=27).value = f'=SUM(AA2:AA{ultima_linha - 1})'
            planilha_principal.cell(row=ultima_linha, column=28).value = f'=SUM(AB2:AB{ultima_linha - 1})'
            planilha_principal.cell(row=ultima_linha, column=29).value = f'=SUM(AC2:AC{ultima_linha - 1})'
            planilha_principal.cell(row=ultima_linha, column=30).value = f'=SUM(AD2:AD{ultima_linha - 1})'
            planilha_principal.cell(row=ultima_linha, column=31).value = f'=SUM(AE2:AE{ultima_linha - 1})'
            planilha_principal.cell(row=ultima_linha, column=32).value = f'=SUM(AF2:AF{ultima_linha - 1})'
            
            planilha_resumo.cell(row=2, column=2).value = f'=SUM((Principal!O{ultima_linha})*(1+E2))' # Material
            planilha_resumo.cell(row=3, column=2).value = f'=SUM((Principal!AA{ultima_linha})*(1+E3))' # Usinagem
            planilha_resumo.cell(row=4, column=2).value = f'=SUM((Principal!W{ultima_linha})*(1+E4))' # Pintura
            planilha_resumo.cell(row=5, column=2).value = f'=SUM((Principal!Z{ultima_linha})*(1+E5))' # Caldeiraria
            planilha_resumo.cell(row=6, column=2).value = f'=SUM((Principal!X{ultima_linha}+Principal!Y{ultima_linha})*(1+E6))' # Tratamentos

            # Salva as alterações no arquivo XLSM
            arquivo_excel.save(arquivo_pcp_zero)
            arquivo_excel.close()

            print('Criação do arquivo PCP concluída!')

        except Exception as e:
            traceback_str = traceback.format_exc()
            stack = traceback.extract_stack()
            function_name = stack[-2].name
            print(f"Ocorreu um erro na função '{function_name}' e o código foi abortado.")
            print("Erro: " + repr(e))
            print("Descrição do erro: " + str(e))
            print("Traceback: " + traceback_str)


# Execução da aplicação
if __name__ == '__main__':
    app = QtW.QApplication(sys.argv)
    UIWindow = UI()
    app.exec_()
